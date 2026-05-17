#!/usr/bin/env python3
"""
Flight Search v2: SFO → Venice area via SerpAPI Google Flights
Searches multiple nearby airports and stop configurations.

Requirements: pip install requests openpyxl
"""

import os
import requests
import json
import sys
import time

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing dependency. Run: pip install openpyxl")
    sys.exit(1)

API_KEY = os.environ.get("SERPAPI_KEY", "")
SEARCH_URL = "https://serpapi.com/search"

# Venice + nearby airports that have better business class connections
DESTINATIONS = [
    ("VCE", "Venice"),
    ("MXP", "Milan Malpensa"),   # ~2.5hr train to Venice
    ("TSF", "Venice Treviso"),
]

RETURN_DATES = ["2026-06-15", "2026-06-16"]
DEPART_DATE = "2026-06-09"

# Try both 1 stop max and 2 stops max
STOP_OPTIONS = [1, 2]

def search_flights(dest_code, return_date, max_stops):
    params = {
        "engine": "google_flights",
        "api_key": API_KEY,
        "departure_id": "SFO",
        "arrival_id": dest_code,
        "outbound_date": DEPART_DATE,
        "return_date": return_date,
        "travel_class": 3,  # business
        "adults": 2,
        "children": 1,
        "stops": max_stops,
        "currency": "USD",
        "hl": "en",
    }
    resp = requests.get(SEARCH_URL, params=params)
    if resp.status_code != 200:
        return None
    data = resp.json()
    if "error" in data and "hasn't returned any results" in data.get("error", ""):
        return None
    return data

def parse_flight(flight, dest_code, dest_name, return_date):
    legs = flight.get("flights", [])
    if not legs:
        return None

    dep = legs[0].get("departure_airport", {})
    arr = legs[-1].get("arrival_airport", {})
    stops = len(legs) - 1
    total_dur = flight.get("total_duration", 0)
    h, m = divmod(total_dur, 60)

    layovers = flight.get("layovers", [])
    layover_strs = []
    for lo in layovers:
        lo_dur = lo.get("duration", 0)
        lo_h, lo_m = divmod(lo_dur, 60)
        layover_strs.append(f"{lo.get('name', '?')} ({lo_h}h {lo_m}m)")

    airlines = ", ".join(flight.get("airlines", []))
    flight_nums = ", ".join(seg.get("flight_number", "?") for seg in legs)

    route_parts = [dep.get("id", "?")]
    for seg in legs:
        route_parts.append(seg.get("arrival_airport", {}).get("id", "?"))

    return {
        "destination": f"{dest_code} ({dest_name})",
        "return_date": return_date,
        "price": flight.get("price", 0),
        "airlines": airlines,
        "flight_numbers": flight_nums,
        "route": " -> ".join(route_parts),
        "stops": stops,
        "total_duration_min": total_dur,
        "duration_str": f"{h}h {m}m",
        "depart_time": dep.get("time", ""),
        "arrive_time": arr.get("time", ""),
        "layovers": " | ".join(layover_strs) if layover_strs else "Nonstop",
    }

def export_to_excel(all_rows, filename="sfo_venice_flights.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Flight Results"

    headers = [
        "Destination", "Return Date", "Price (USD)", "Airlines", "Flight #s",
        "Route", "Stops", "Total Duration", "Depart Time", "Arrive Time",
        "Layover Details"
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin_border = Border(bottom=Side(style="thin", color="D9D9D9"))

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    all_rows.sort(key=lambda x: x["price"])

    for row_idx, flight in enumerate(all_rows, 2):
        values = [
            flight["destination"],
            flight["return_date"],
            flight["price"],
            flight["airlines"],
            flight["flight_numbers"],
            flight["route"],
            flight["stops"],
            flight["duration_str"],
            flight["depart_time"],
            flight["arrive_time"],
            flight["layovers"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

        ws.cell(row=row_idx, column=3).number_format = '$#,##0'

    widths = [20, 12, 14, 22, 18, 28, 8, 16, 14, 14, 35]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    wb.save(filename)
    return filename

def main():
    all_rows = []
    search_count = 0

    for dest_code, dest_name in DESTINATIONS:
        for ret_date in RETURN_DATES:
            for max_stops in STOP_OPTIONS:
                label = f"SFO -> {dest_code} | return {ret_date} | max {max_stops} stop(s)"
                print(f"  Searching: {label}...")

                data = search_flights(dest_code, ret_date, max_stops)
                search_count += 1

                if not data:
                    print(f"    No results")
                    time.sleep(0.5)  # rate limit courtesy
                    continue

                best = data.get("best_flights", [])
                other = data.get("other_flights", [])
                count = 0

                for flight in best + other:
                    parsed = parse_flight(flight, dest_code, dest_name, ret_date)
                    if parsed:
                        all_rows.append(parsed)
                        count += 1

                print(f"    Found {count} options")

                # Price insights
                insights = data.get("price_insights", {})
                if insights.get("lowest_price"):
                    print(f"    Lowest price: ${insights['lowest_price']:,}")

                time.sleep(0.5)

    print(f"\n{'='*60}")
    print(f"  TOTAL: {len(all_rows)} options across {search_count} searches")
    print(f"{'='*60}")

    if all_rows:
        # Deduplicate by flight numbers + return date
        seen = set()
        unique = []
        for r in all_rows:
            key = (r["flight_numbers"], r["return_date"])
            if key not in seen:
                seen.add(key)
                unique.append(r)
        all_rows = unique
        print(f"  {len(all_rows)} unique options after dedup\n")

        by_price = sorted(all_rows, key=lambda x: x["price"])
        print(f"  -- Top 5 Cheapest --\n")
        for i, r in enumerate(by_price[:5], 1):
            print(f"  #{i}  ${r['price']:,}  |  {r['duration_str']}  |  {r['airlines']}")
            print(f"       {r['route']}  |  Dest: {r['destination']}  |  Return: {r['return_date']}")
            print(f"       {r['layovers']}")
            print()

        by_time = sorted(all_rows, key=lambda x: x["total_duration_min"])
        print(f"  -- Top 5 Fastest --\n")
        for i, r in enumerate(by_time[:5], 1):
            print(f"  #{i}  {r['duration_str']}  |  ${r['price']:,}  |  {r['airlines']}")
            print(f"       {r['route']}  |  Dest: {r['destination']}  |  Return: {r['return_date']}")
            print(f"       {r['layovers']}")
            print()

        fname = export_to_excel(all_rows)
        print(f"Spreadsheet saved: {fname}")
    else:
        print("\n  No flights found across any search.")
        print("  This likely means business class inventory isn't loaded yet for June 2026.")
        print("  Try checking Google Flights directly in a browser to confirm.")

if __name__ == "__main__":
    main()
